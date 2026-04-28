from fastapi import FastAPI, APIRouter, HTTPException, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
import os
import sqlite3
import json
import uuid as uuid_lib
import re
import shutil
import unicodedata
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
import logging
from datetime import datetime
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

DB_PATH = ROOT_DIR / "catalog.db"
FRONTEND_PUBLIC_DIR = ROOT_DIR.parent / "frontend" / "public"
IMAGE_UPLOAD_DIR = FRONTEND_PUBLIC_DIR / "images" / "uploads"
IS_DEVELOPMENT = os.environ.get('IS_DEVELOPMENT', 'true').lower() == 'true'

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
IMPORT_METADATA_PREFIX = "__IMPORT_METADATA__"

# SQLite helpers
def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def row_to_dict(row):
    if row is None:
        return None
    d = dict(row)
    if 'ativo' in d:
        d['ativo'] = bool(d['ativo'])
    return d

def slugify(value: str) -> str:
    value = re.sub(r'[^a-zA-Z0-9]+', '-', value.strip().lower())
    return value.strip('-') or 'item'

def normalize_catalog_text(value: str) -> str:
    normalized = unicodedata.normalize('NFD', str(value or ''))
    normalized = ''.join(char for char in normalized if unicodedata.category(char) != 'Mn')
    normalized = re.sub(r'\s+', ' ', normalized).strip().lower()
    return normalized

def parse_cash_observation(value: Optional[str]):
    raw = str(value or '')
    if not raw.startswith(IMPORT_METADATA_PREFIX):
        return None
    newline_index = raw.find('\n')
    metadata_text = raw[len(IMPORT_METADATA_PREFIX):] if newline_index < 0 else raw[len(IMPORT_METADATA_PREFIX):newline_index]
    try:
        return json.loads(metadata_text)
    except json.JSONDecodeError:
        return None

def get_catalog_sales_ranking(conn):
    products = [
        row_to_dict(row)
        for row in conn.execute(
            "SELECT * FROM products WHERE ativo = 1 AND categoria != 'Bebidas'"
        ).fetchall()
    ]
    catalog_items = products
    sales_by_slug = {}

    cash_rows = conn.execute(
        "SELECT observacao FROM cash_entries WHERE descricao IN ('Vendas WhatsApp', 'Vendas iFood')"
    ).fetchall()

    for row in cash_rows:
        metadata = parse_cash_observation(row['observacao'])
        if not metadata:
            continue
        for batch in metadata.get('batches', []):
            for order in batch.get('orders', []):
                for item in order.get('items', []):
                    item_name = item.get('name')
                    item_quantity = int(item.get('quantity') or 0)
                    if not item_name or item_quantity <= 0:
                        continue
                    slug = normalize_catalog_text(item_name)
                    sales_by_slug[slug] = sales_by_slug.get(slug, 0) + item_quantity

    ranked_items = []
    for item in catalog_items:
        sold_count = sales_by_slug.get(normalize_catalog_text(item.get('nome')), 0)
        ranked_items.append({
            **item,
            'sold_count': sold_count,
        })

    ranked_items.sort(key=lambda item: (-item['sold_count'], item['nome']))
    return ranked_items

def parse_ifood_datetime(value):
    if isinstance(value, datetime):
        return value
    text = str(value or '').strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None

def parse_ifood_report(file_obj):
    workbook = load_workbook(file_obj, read_only=False, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []

    headers = {str(value or '').strip(): index for index, value in enumerate(header_row)}

    def cell(row, column_name):
        index = headers.get(column_name)
        if index is None or index >= len(row):
            return None
        return row[index]

    orders = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        full_id = cell(row, 'ID COMPLETO DO PEDIDO')
        short_id = cell(row, 'ID CURTO DO PEDIDO')
        status = str(cell(row, 'STATUS FINAL DO PEDIDO') or '').strip().upper()
        net = float(cell(row, 'VALOR LIQUIDO (R$)') or 0)

        if not full_id or net <= 0 or status not in {'CONCLUIDO', 'CONCLUÍDO'}:
            continue

        order_datetime = parse_ifood_datetime(cell(row, 'DATA E HORA DO PEDIDO'))
        parsed_date = order_datetime.strftime('%Y-%m-%d') if order_datetime else ''
        created_at = order_datetime.strftime('%d/%m/%Y %H:%M') if order_datetime else ''
        time = order_datetime.strftime('%H:%M') if order_datetime else ''

        orders.append({
            'id': str(full_id).strip(),
            'orderId': str(short_id).strip(),
            'createdAt': created_at,
            'parsedDate': parsed_date,
            'time': time,
            'gross': float(cell(row, 'TOTAL PAGO PELO CLIENTE (R$)') or cell(row, 'VALOR DOS ITENS (R$)') or 0),
            'net': net,
            'status': status,
            'paymentMethod': str(cell(row, 'FORMA DE PAGAMENTO') or '').strip(),
            'deliveryType': str(cell(row, 'TIPO DE ENTREGA') or '').strip(),
            'channel': str(cell(row, 'CANAL DE VENDA') or '').strip(),
        })

    return orders

def save_uploaded_image(file: UploadFile, scope: str, item_id: Optional[str] = None) -> str:
    suffix = Path(file.filename or '').suffix.lower()
    allowed_extensions = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}
    if suffix not in allowed_extensions:
        raise ValueError("Formato de imagem invalido. Use JPG, JPEG, PNG, WEBP ou GIF.")

    IMAGE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    scope_slug = slugify(scope)
    item_slug = slugify(item_id or str(uuid_lib.uuid4()))
    prefix = f"{scope_slug}-{item_slug}"

    for existing_file in IMAGE_UPLOAD_DIR.glob(f"{prefix}.*"):
        existing_file.unlink(missing_ok=True)

    destination = IMAGE_UPLOAD_DIR / f"{prefix}{suffix}"
    with destination.open('wb') as buffer:
        shutil.copyfileobj(file.file, buffer)

    return f"/images/uploads/{destination.name}"

# Pydantic models
class ProductCreate(BaseModel):
    nome: str
    descricao: Optional[str] = None
    preco: float
    desconto: Optional[float] = 0
    foto: Optional[str] = None
    categoria: str
    vendas: Optional[int] = 0
    ativo: Optional[bool] = True

class ProductUpdate(BaseModel):
    nome: Optional[str] = None
    descricao: Optional[str] = None
    preco: Optional[float] = None
    desconto: Optional[float] = None
    foto: Optional[str] = None
    categoria: Optional[str] = None
    vendas: Optional[int] = None
    ativo: Optional[bool] = None

class ComboCreate(BaseModel):
    nome: str
    descricao: Optional[str] = None
    produto_ids: Optional[List[str]] = []
    valor: float
    desconto: Optional[float] = 0
    vendas: Optional[int] = 0
    foto: Optional[str] = None
    ativo: Optional[bool] = True

class ComboUpdate(BaseModel):
    nome: Optional[str] = None
    descricao: Optional[str] = None
    produto_ids: Optional[List[str]] = None
    valor: Optional[float] = None
    desconto: Optional[float] = None
    vendas: Optional[int] = None
    foto: Optional[str] = None
    ativo: Optional[bool] = None

class PaymentMethodCreate(BaseModel):
    nome: str
    ativo: Optional[bool] = True

class PaymentMethodUpdate(BaseModel):
    nome: Optional[str] = None
    ativo: Optional[bool] = None

class AddonCreate(BaseModel):
    nome: str
    preco: float
    ativo: Optional[bool] = True

class AddonUpdate(BaseModel):
    nome: Optional[str] = None
    preco: Optional[float] = None
    ativo: Optional[bool] = None

class StoreSettingsUpdate(BaseModel):
    whatsapp: Optional[str] = None
    delivery_time: Optional[str] = None
    business_hours: Optional[str] = None
    promotion_product_uuid: Optional[str] = None
    promotion_price: Optional[float] = None
    promotion_active: Optional[bool] = None
    receivable_reset_day: Optional[int] = None

class CashEntryCreate(BaseModel):
    tipo: str
    categoria: str
    descricao: str
    valor: float
    forma_pagamento: Optional[str] = None
    data_lancamento: Optional[str] = None
    observacao: Optional[str] = None

class CashEntryUpdate(BaseModel):
    tipo: Optional[str] = None
    categoria: Optional[str] = None
    descricao: Optional[str] = None
    valor: Optional[float] = None
    forma_pagamento: Optional[str] = None
    data_lancamento: Optional[str] = None
    observacao: Optional[str] = None

class ReceivableWithdrawalCreate(BaseModel):
    valor: float
    data_saque: Optional[str] = None
    observacao: Optional[str] = None

class StockItemCreate(BaseModel):
    nome: str
    unidade: str
    quantidade: float
    valor_pago: float
    ativo: Optional[bool] = True

class StockItemUpdate(BaseModel):
    nome: Optional[str] = None
    unidade: Optional[str] = None
    quantidade: Optional[float] = None
    valor_pago: Optional[float] = None
    ativo: Optional[bool] = None

class ProductRecipeCreate(BaseModel):
    product_uuid: str
    stock_item_uuid: str
    quantidade_utilizada: float

class ProductRecipeUpdate(BaseModel):
    product_uuid: Optional[str] = None
    stock_item_uuid: Optional[str] = None
    quantidade_utilizada: Optional[float] = None

# Database initialization
def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS products (
        uuid TEXT PRIMARY KEY,
        nome TEXT NOT NULL,
        descricao TEXT,
        preco REAL NOT NULL,
        desconto REAL DEFAULT 0,
        foto TEXT,
        categoria TEXT NOT NULL,
        vendas INTEGER DEFAULT 0,
        ativo INTEGER DEFAULT 1
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS combos (
        uuid TEXT PRIMARY KEY,
        nome TEXT NOT NULL,
        descricao TEXT,
        produto_ids TEXT DEFAULT '[]',
        valor REAL NOT NULL,
        desconto REAL DEFAULT 0,
        vendas INTEGER DEFAULT 0,
        foto TEXT,
        ativo INTEGER DEFAULT 1
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS payment_methods (
        uuid TEXT PRIMARY KEY,
        nome TEXT NOT NULL,
        ativo INTEGER DEFAULT 1
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS addons (
        uuid TEXT PRIMARY KEY,
        nome TEXT NOT NULL,
        preco REAL NOT NULL,
        ativo INTEGER DEFAULT 1
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS store_settings (
        id INTEGER PRIMARY KEY CHECK (id = 1),
        whatsapp TEXT NOT NULL,
        delivery_time TEXT NOT NULL,
        business_hours TEXT NOT NULL,
        promotion_product_uuid TEXT,
        promotion_price REAL,
        promotion_active INTEGER DEFAULT 0,
        receivable_reset_day INTEGER DEFAULT 15
    )''')
    existing_columns = {
        row['name']
        for row in cur.execute("PRAGMA table_info(store_settings)").fetchall()
    }
    if 'promotion_product_uuid' not in existing_columns:
        cur.execute("ALTER TABLE store_settings ADD COLUMN promotion_product_uuid TEXT")
    if 'promotion_price' not in existing_columns:
        cur.execute("ALTER TABLE store_settings ADD COLUMN promotion_price REAL")
    if 'promotion_active' not in existing_columns:
        cur.execute("ALTER TABLE store_settings ADD COLUMN promotion_active INTEGER DEFAULT 0")
    if 'receivable_reset_day' not in existing_columns:
        cur.execute("ALTER TABLE store_settings ADD COLUMN receivable_reset_day INTEGER DEFAULT 15")
    cur.execute('''CREATE TABLE IF NOT EXISTS cash_entries (
        uuid TEXT PRIMARY KEY,
        tipo TEXT NOT NULL,
        categoria TEXT NOT NULL,
        descricao TEXT NOT NULL,
        valor REAL NOT NULL,
        forma_pagamento TEXT,
        data_lancamento TEXT NOT NULL,
        observacao TEXT,
        created_at TEXT NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS stock_items (
        uuid TEXT PRIMARY KEY,
        nome TEXT NOT NULL,
        unidade TEXT NOT NULL,
        quantidade REAL NOT NULL,
        valor_pago REAL NOT NULL,
        ativo INTEGER DEFAULT 1
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS product_recipes (
        uuid TEXT PRIMARY KEY,
        product_uuid TEXT NOT NULL,
        stock_item_uuid TEXT NOT NULL,
        quantidade_utilizada REAL NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS receivable_withdrawals (
        uuid TEXT PRIMARY KEY,
        valor REAL NOT NULL,
        data_saque TEXT NOT NULL,
        observacao TEXT,
        created_at TEXT NOT NULL
    )''')
    conn.commit()
    cur.execute("SELECT COUNT(*) as cnt FROM products")
    if cur.fetchone()['cnt'] == 0:
        seed_data(conn)
    cur.execute("SELECT COUNT(*) as cnt FROM payment_methods")
    if cur.fetchone()['cnt'] == 0:
        seed_payment_methods(conn)
    cur.execute("SELECT COUNT(*) as cnt FROM addons")
    if cur.fetchone()['cnt'] == 0:
        seed_addons(conn)
    cur.execute("SELECT COUNT(*) as cnt FROM store_settings")
    if cur.fetchone()['cnt'] == 0:
        seed_store_settings(conn)
    conn.close()

def seed_payment_methods(conn):
    cur = conn.cursor()
    methods = [
        ("pm-001", "Pix", 1),
        ("pm-002", "Pagamento na entrega Crédito", 1),
        ("pm-003", "Pagamento na entrega Débito", 1),
        ("pm-004", "Dinheiro", 1),
    ]
    for m in methods:
        cur.execute("INSERT INTO payment_methods VALUES (?,?,?)", m)
    conn.commit()
    logger.info("Payment methods seeded")

def seed_addons(conn):
    cur = conn.cursor()
    addons = [
        ("addon-001", "Doritos", 3.00, 1),
        ("addon-002", "Mussarela", 3.00, 1),
        ("addon-003", "Catupiry", 3.00, 1),
        ("addon-004", "Cheddar", 3.00, 1),
    ]
    for addon in addons:
        cur.execute("INSERT INTO addons VALUES (?,?,?,?)", addon)
    conn.commit()
    logger.info("Addons seeded")

def seed_store_settings(conn):
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO store_settings (
            id, whatsapp, delivery_time, business_hours,
            promotion_product_uuid, promotion_price, promotion_active, receivable_reset_day
        ) VALUES (1, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "5535998160726",
            "40 min",
            "Seg a Sex: 18:00 - 23:00\nSab e Dom: 17:00 - 00:00",
            "prod-001",
            24.90,
            1,
            15,
        ),
    )
    conn.commit()
    logger.info("Store settings seeded")

def seed_data(conn):
    cur = conn.cursor()
    products = [
        ("prod-001", "Batata Recheada De Carne Com Cheddar E Doritos", "Batata recheada feita na hora, bem caprichada. Carne moida bem temperada com molho de tomate da casa cremoso envolvendo tudo, feito com carne, molho de tomate da casa, na finalizacao vai mussarela, cheddar nas bordas e cheiro verde. Aqui a gente nao economiza nao, e recheio generoso pra comer feliz do comeco ao fim.", 28.90, 10, "/images/products/batata-carne-cheddar-doritos.jpeg", "Batatas Recheadas 400g", 0, 1),
        ("prod-002", "Batata Recheada De Strogonoff De Frango", "Batata recheada feita na hora, com um sabor excepcional. Recheio cremoso de strogonoff de frango feito com frango, cebola, molho de tomate da casa, creme de leite, na finalizacao vai mussarela, catupiry nas bordas, batata palha crocante e cheiro verde pra finalizar. E pode ficar tranquilo: aqui e sem miseria mesmo, viu? E recheio de verdade!", 23.90, 10, "/images/products/batata-strogonoff-frango.jpeg", "Batatas Recheadas 400g", 0, 1),
        ("prod-003", "Batata Recheada De Frango Com Catupiry", "Batata recheada feita na hora, bem caprichada e daquele jeito que da orgulho de servir. Frango desfiado bem temperado com muito catupiry cremoso envolvendo tudo, feito com frango, molho de tomate da casa e catupiry cremoso, na finalizacao vai mussarela, catupiry nas bordas, bacon e cheiro verde. Aqui a gente nao economiza nao, e recheio generoso pra comer feliz do comeco ao fim.", 25.90, 10, "/images/products/batata-frango-catupiry.jpeg", "Batatas Recheadas 400g", 0, 1),
        ("prod-004", "Batata Recheada De Sabor Pizza", "Batata quentinha, feita na hora e carregada de sabor. Presunto e queijo super cremoso, tomate, finalizacao com mussarela, Doritos e cream cheese. E aquela batata forte, bem servida e sem miseria, pra quem gosta de comer de verdade.", 28.90, 10, "/images/products/batata-sabor-pizza.jpeg", "Batatas Recheadas 400g", 0, 1),
        ("prod-005", "Batata Recheada De Brocolis Com Queijo", "Batata recheada feita na hora, bem caprichada e cheia de sabor. Brocolis temperadinho com queijo cremoso derretendo, combinacao que voce nem sabia que precisava. Mais leve? Talvez. Mas sem miseria? Nunca.", 22.90, 10, "/images/products/batata-brocolis-queijo.jpeg", "Batatas Recheadas 400g", 1, 1),
        ("prod-006", "Coca-Cola Lata 350ml", "Refrigerante Coca-Cola tradicional em lata de 350ml, gelado e pronto para acompanhar seu pedido.", 7.00, 0, None, "Bebidas", 0, 1),
        ("prod-007", "Coca-Cola Zero Lata 350ml", "Refrigerante Coca-Cola Zero em lata de 350ml, sem acucar e com o sabor classico que combina com qualquer refeicao.", 7.00, 0, None, "Bebidas", 0, 1),
        ("prod-008", "Guarana Antarctica Lata 350ml", "Refrigerante Guarana Antarctica tradicional em lata de 350ml, gelado e ideal para acompanhar seu lanche.", 7.00, 0, None, "Bebidas", 0, 1),
        ("prod-009", "Guarana Antarctica Zero Lata 350ml", "Refrigerante Guarana Antarctica Zero em lata de 350ml, sem acucar e com o sabor leve e refrescante.", 7.00, 0, None, "Bebidas", 0, 1),
    ]
    for product in products:
        cur.execute("INSERT INTO products VALUES (?,?,?,?,?,?,?,?,?)", product)
    conn.commit()
    logger.info("Database seeded with curated potato menu")
    return
    products = [
        ("prod-001", "Batata Carne Seca", "Batata recheada com carne seca desfiada, cheddar cremoso, cebola caramelizada e cebolinha", 32.90, 0, "https://images.unsplash.com/photo-1645673197548-9adfa2ea55dc?w=400&h=300&fit=crop", "Batatas Recheadas", 200, 1),
        ("prod-002", "Batata Frango Catupiry", "Frango desfiado temperado com catupiry original e orégano", 29.90, 10, "https://images.unsplash.com/photo-1639744210631-209fce3e256c?w=400&h=300&fit=crop", "Batatas Recheadas", 180, 1),
        ("prod-003", "Batata Calabresa", "Calabresa acebolada com queijo mussarela derretido e molho especial", 28.90, 0, "https://images.unsplash.com/photo-1761712826074-5f1bab2b2f32?w=400&h=300&fit=crop", "Batatas Recheadas", 150, 1),
        ("prod-004", "Batata 4 Queijos", "Mussarela, cheddar, provolone e parmesão com creme de leite", 34.90, 15, "https://images.unsplash.com/photo-1645673197548-9adfa2ea55dc?w=400&h=300&fit=crop&crop=center", "Batatas Recheadas", 130, 1),
        ("prod-005", "Batata Bacon Cheddar", "Bacon crocante picado com cheddar cremoso derretido e cebolinha", 31.90, 0, "https://images.unsplash.com/photo-1639744210631-209fce3e256c?w=400&h=300&fit=crop&crop=top", "Batatas Recheadas", 170, 1),
        ("prod-006", "Batata Brócolis com Cheddar", "Brócolis cozido no vapor com cheddar cremoso e bacon bits", 30.90, 0, "https://images.unsplash.com/photo-1761712826074-5f1bab2b2f32?w=400&h=300&fit=crop&crop=center", "Batatas Recheadas", 90, 1),
        ("prod-007", "Batata Strogonoff", "Strogonoff de carne cremoso com batata palha e salsinha", 33.90, 10, "https://images.unsplash.com/photo-1645673197548-9adfa2ea55dc?w=400&h=300&fit=crop&crop=top", "Batatas Recheadas", 160, 1),
        ("prod-008", "Batata Vegana", "Cogumelos refogados, ervilha, tomate seco e azeite trufado", 27.90, 0, "https://images.unsplash.com/photo-1639744210631-209fce3e256c?w=400&h=300&fit=crop&crop=center", "Batatas Recheadas", 60, 1),
        ("prod-009", "Refrigerante 2 Litros", "Coca-Cola, Guaraná ou Fanta - escolha na observação", 14.00, 0, "https://images.unsplash.com/photo-1632996988606-274cfd06eb68?w=400&h=300&fit=crop", "Bebidas", 300, 1),
        ("prod-010", "Suco Natural 500ml", "Laranja, limão, maracujá ou abacaxi - feito na hora", 12.00, 0, "https://images.pexels.com/photos/4113669/pexels-photo-4113669.jpeg?auto=compress&cs=tinysrgb&w=400", "Bebidas", 80, 1),
        ("prod-011", "Água Mineral 500ml", "Água mineral sem gás gelada", 5.00, 0, "https://images.unsplash.com/photo-1653122024753-d7322d90b3d7?w=400&h=300&fit=crop", "Bebidas", 250, 1),
        ("prod-012", "Refrigerante Lata 350ml", "Coca-Cola, Guaraná, Fanta ou Sprite gelada", 7.00, 0, "https://images.pexels.com/photos/31336111/pexels-photo-31336111.jpeg?auto=compress&cs=tinysrgb&w=400", "Bebidas", 350, 1),
        ("prod-013", "Petit Gateau", "Bolo quente de chocolate com sorvete de creme e calda especial", 24.90, 20, "https://images.unsplash.com/photo-1551024601-bec78aea704b?w=400&h=300&fit=crop", "Sobremesas", 70, 1),
        ("prod-014", "Brownie com Sorvete", "Brownie artesanal quentinho com sorvete de creme e calda", 22.90, 0, "https://images.unsplash.com/photo-1551024506-0bccd828d307?w=400&h=300&fit=crop", "Sobremesas", 60, 1),
    ]
    for p in products:
        cur.execute("INSERT INTO products VALUES (?,?,?,?,?,?,?,?,?)", p)
    combos = [
        ("combo-001", "Combo Recheiaê Família", "Batata Carne Seca + Batata Frango Catupiry + Refrigerante 2L", json.dumps(["prod-001", "prod-002", "prod-009"]), 69.90, 10, 150, "https://images.unsplash.com/photo-1645673197548-9adfa2ea55dc?w=400&h=300&fit=crop", 1),
        ("combo-002", "Combo Casal", "Batata Bacon Cheddar + Batata Calabresa + 2 Refri Lata", json.dumps(["prod-005", "prod-003", "prod-012"]), 59.90, 5, 100, "https://images.unsplash.com/photo-1639744210631-209fce3e256c?w=400&h=300&fit=crop", 1),
        ("combo-003", "Combo Solo", "Batata Strogonoff + Refrigerante Lata + Brownie", json.dumps(["prod-007", "prod-012", "prod-014"]), 54.90, 0, 80, "https://images.unsplash.com/photo-1761712826074-5f1bab2b2f32?w=400&h=300&fit=crop", 1),
    ]
    for c in combos:
        cur.execute("INSERT INTO combos VALUES (?,?,?,?,?,?,?,?,?)", c)
    conn.commit()
    logger.info("Database seeded with sample data")

# Public routes
@api_router.get("/")
def root():
    return {"message": "Recheiaê API"}

@api_router.get("/products")
def get_products():
    conn = get_db()
    rows = conn.execute("SELECT * FROM products WHERE ativo = 1 ORDER BY vendas DESC").fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.get("/products/all")
def get_all_products():
    conn = get_db()
    rows = conn.execute("SELECT * FROM products ORDER BY categoria, nome").fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.get("/combos")
def get_combos():
    conn = get_db()
    rows = conn.execute("SELECT * FROM combos WHERE ativo = 1 ORDER BY vendas DESC").fetchall()
    conn.close()
    result = []
    for r in rows:
        d = row_to_dict(r)
        d['produto_ids'] = json.loads(d.get('produto_ids') or '[]')
        result.append(d)
    return result

@api_router.get("/combos/all")
def get_all_combos():
    conn = get_db()
    rows = conn.execute("SELECT * FROM combos ORDER BY nome").fetchall()
    conn.close()
    result = []
    for r in rows:
        d = row_to_dict(r)
        d['produto_ids'] = json.loads(d.get('produto_ids') or '[]')
        result.append(d)
    return result

@api_router.get("/categories")
def get_categories():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT categoria FROM products WHERE ativo = 1 ORDER BY categoria").fetchall()
    conn.close()
    return [r['categoria'] for r in rows]

@api_router.get("/products/top")
def get_top_products():
    conn = get_db()
    ranking = get_catalog_sales_ranking(conn)
    conn.close()
    return [item for item in ranking if item.get('sold_count', 0) > 0][:3]

@api_router.get("/payment-methods")
def get_payment_methods():
    conn = get_db()
    rows = conn.execute("SELECT * FROM payment_methods WHERE ativo = 1 ORDER BY nome").fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.get("/payment-methods/all")
def get_all_payment_methods():
    conn = get_db()
    rows = conn.execute("SELECT * FROM payment_methods ORDER BY nome").fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.get("/addons")
def get_addons():
    conn = get_db()
    rows = conn.execute("SELECT * FROM addons WHERE ativo = 1 ORDER BY nome").fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.get("/addons/all")
def get_all_addons():
    conn = get_db()
    rows = conn.execute("SELECT * FROM addons ORDER BY nome").fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.get("/settings")
def get_store_settings():
    conn = get_db()
    row = conn.execute(
        """
        SELECT
            whatsapp,
            delivery_time,
            business_hours,
            promotion_product_uuid,
            promotion_price,
            promotion_active,
            receivable_reset_day
        FROM store_settings
        WHERE id = 1
        """
    ).fetchone()
    conn.close()
    if not row:
        return {
            "whatsapp": "5535998160726",
            "delivery_time": "40 min",
            "business_hours": "Seg a Sex: 18:00 - 23:00\nSab e Dom: 17:00 - 00:00",
            "promotion_product_uuid": "prod-001",
            "promotion_price": 24.90,
            "promotion_active": 1,
            "receivable_reset_day": 15,
        }
    return dict(row)

# Admin routes
@api_router.post("/admin/uploads/image")
async def upload_admin_image(
    file: UploadFile = File(...),
    scope: str = Form("product"),
    item_id: Optional[str] = Form(None),
):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")

    try:
        image_path = save_uploaded_image(file, scope=scope, item_id=item_id)
        return {"path": image_path}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        await file.close()

@api_router.post("/admin/products")
def create_product(product: ProductCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    conn.execute(
        "INSERT INTO products (uuid, nome, descricao, preco, desconto, foto, categoria, vendas, ativo) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (new_uuid, product.nome, product.descricao, product.preco, product.desconto or 0, product.foto, product.categoria, product.vendas or 0, 1 if product.ativo else 0)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM products WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.put("/admin/products/{product_uuid}")
def update_product(product_uuid: str, product: ProductUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT * FROM products WHERE uuid = ?", (product_uuid,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Product not found")
    updates = {}
    for field, value in product.model_dump(exclude_unset=True).items():
        if field == 'ativo':
            updates[field] = 1 if value else 0
        else:
            updates[field] = value
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(f"UPDATE products SET {set_clause} WHERE uuid = ?", list(updates.values()) + [product_uuid])
        conn.commit()
    row = conn.execute("SELECT * FROM products WHERE uuid = ?", (product_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.delete("/admin/products/{product_uuid}")
def delete_product(product_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM products WHERE uuid = ?", (product_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Product deleted"}

@api_router.post("/admin/combos")
def create_combo(combo: ComboCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    conn.execute(
        "INSERT INTO combos (uuid, nome, descricao, produto_ids, valor, desconto, vendas, foto, ativo) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (new_uuid, combo.nome, combo.descricao, json.dumps(combo.produto_ids or []), combo.valor, combo.desconto or 0, combo.vendas or 0, combo.foto, 1 if combo.ativo else 0)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM combos WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    d = row_to_dict(row)
    d['produto_ids'] = json.loads(d.get('produto_ids') or '[]')
    return d

@api_router.put("/admin/combos/{combo_uuid}")
def update_combo(combo_uuid: str, combo: ComboUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT * FROM combos WHERE uuid = ?", (combo_uuid,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Combo not found")
    updates = {}
    for field, value in combo.model_dump(exclude_unset=True).items():
        if field == 'ativo':
            updates[field] = 1 if value else 0
        elif field == 'produto_ids':
            updates[field] = json.dumps(value or [])
        else:
            updates[field] = value
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(f"UPDATE combos SET {set_clause} WHERE uuid = ?", list(updates.values()) + [combo_uuid])
        conn.commit()
    row = conn.execute("SELECT * FROM combos WHERE uuid = ?", (combo_uuid,)).fetchone()
    conn.close()
    d = row_to_dict(row)
    d['produto_ids'] = json.loads(d.get('produto_ids') or '[]')
    return d

@api_router.delete("/admin/combos/{combo_uuid}")
def delete_combo(combo_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM combos WHERE uuid = ?", (combo_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Combo deleted"}

# Payment methods admin
@api_router.post("/admin/payment-methods")
def create_payment_method(pm: PaymentMethodCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    conn.execute("INSERT INTO payment_methods (uuid, nome, ativo) VALUES (?, ?, ?)",
                 (new_uuid, pm.nome, 1 if pm.ativo else 0))
    conn.commit()
    row = conn.execute("SELECT * FROM payment_methods WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.put("/admin/payment-methods/{pm_uuid}")
def update_payment_method(pm_uuid: str, pm: PaymentMethodUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT * FROM payment_methods WHERE uuid = ?", (pm_uuid,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Payment method not found")
    updates = {}
    for field, value in pm.model_dump(exclude_unset=True).items():
        if field == 'ativo':
            updates[field] = 1 if value else 0
        else:
            updates[field] = value
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(f"UPDATE payment_methods SET {set_clause} WHERE uuid = ?", list(updates.values()) + [pm_uuid])
        conn.commit()
    row = conn.execute("SELECT * FROM payment_methods WHERE uuid = ?", (pm_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.delete("/admin/payment-methods/{pm_uuid}")
def delete_payment_method(pm_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM payment_methods WHERE uuid = ?", (pm_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Payment method deleted"}

@api_router.post("/admin/addons")
def create_addon(addon: AddonCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    conn.execute(
        "INSERT INTO addons (uuid, nome, preco, ativo) VALUES (?, ?, ?, ?)",
        (new_uuid, addon.nome, addon.preco, 1 if addon.ativo else 0),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM addons WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.put("/admin/addons/{addon_uuid}")
def update_addon(addon_uuid: str, addon: AddonUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT * FROM addons WHERE uuid = ?", (addon_uuid,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Addon not found")
    updates = {}
    for field, value in addon.model_dump(exclude_unset=True).items():
        if field == 'ativo':
            updates[field] = 1 if value else 0
        else:
            updates[field] = value
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(f"UPDATE addons SET {set_clause} WHERE uuid = ?", list(updates.values()) + [addon_uuid])
        conn.commit()
    row = conn.execute("SELECT * FROM addons WHERE uuid = ?", (addon_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.delete("/admin/addons/{addon_uuid}")
def delete_addon(addon_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM addons WHERE uuid = ?", (addon_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Addon deleted"}

@api_router.put("/admin/settings")
def update_store_settings(settings: StoreSettingsUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT id FROM store_settings WHERE id = 1").fetchone()
    payload = {
        "whatsapp": "5535998160726",
        "delivery_time": "40 min",
        "business_hours": "Seg a Sex: 18:00 - 23:00\nSab e Dom: 17:00 - 00:00",
        "promotion_product_uuid": None,
        "promotion_price": None,
        "promotion_active": 0,
        "receivable_reset_day": 15,
    }
    if existing:
        row = conn.execute(
            """
            SELECT
                whatsapp,
                delivery_time,
                business_hours,
                promotion_product_uuid,
                promotion_price,
                promotion_active,
                receivable_reset_day
            FROM store_settings
            WHERE id = 1
            """
        ).fetchone()
        payload.update(dict(row))
    payload.update(settings.model_dump(exclude_unset=True))
    if existing:
        conn.execute(
            """
            UPDATE store_settings
            SET
                whatsapp = ?,
                delivery_time = ?,
                business_hours = ?,
                promotion_product_uuid = ?,
                promotion_price = ?,
                promotion_active = ?,
                receivable_reset_day = ?
            WHERE id = 1
            """,
            (
                payload["whatsapp"],
                payload["delivery_time"],
                payload["business_hours"],
                payload["promotion_product_uuid"],
                payload["promotion_price"],
                1 if payload["promotion_active"] else 0,
                payload["receivable_reset_day"] or 15,
            ),
        )
    else:
        conn.execute(
            """
            INSERT INTO store_settings (
                id, whatsapp, delivery_time, business_hours,
                promotion_product_uuid, promotion_price, promotion_active, receivable_reset_day
            ) VALUES (1, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload["whatsapp"],
                payload["delivery_time"],
                payload["business_hours"],
                payload["promotion_product_uuid"],
                payload["promotion_price"],
                1 if payload["promotion_active"] else 0,
                payload["receivable_reset_day"] or 15,
            ),
        )
    conn.commit()
    row = conn.execute(
        """
        SELECT
            whatsapp,
            delivery_time,
            business_hours,
            promotion_product_uuid,
            promotion_price,
            promotion_active,
            receivable_reset_day
        FROM store_settings
        WHERE id = 1
        """
    ).fetchone()
    conn.close()
    return dict(row)

@api_router.get("/admin/cash-entries")
def get_cash_entries():
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM cash_entries ORDER BY data_lancamento DESC, created_at DESC"
    ).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.get("/admin/receivable-withdrawals")
def get_receivable_withdrawals():
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM receivable_withdrawals ORDER BY data_saque DESC, created_at DESC"
    ).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@api_router.post("/admin/receivable-withdrawals")
def create_receivable_withdrawal(withdrawal: ReceivableWithdrawalCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    now_iso = datetime.now().isoformat(timespec='seconds')
    data_saque = withdrawal.data_saque or now_iso[:10]
    conn.execute(
        """
        INSERT INTO receivable_withdrawals (
            uuid, valor, data_saque, observacao, created_at
        ) VALUES (?, ?, ?, ?, ?)
        """,
        (
            new_uuid,
            withdrawal.valor,
            data_saque,
            withdrawal.observacao,
            now_iso,
        ),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM receivable_withdrawals WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.delete("/admin/receivable-withdrawals/{withdrawal_uuid}")
def delete_receivable_withdrawal(withdrawal_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM receivable_withdrawals WHERE uuid = ?", (withdrawal_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Receivable withdrawal deleted"}

@api_router.post("/admin/imports/ifood-report")
async def import_ifood_report(file: UploadFile = File(...)):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")

    filename = (file.filename or '').lower()
    if not filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx do iFood")

    try:
        file.file.seek(0)
        orders = parse_ifood_report(file.file)
        return {'orders': orders}
    except Exception as error:
        logger.exception("Failed to parse iFood report")
        raise HTTPException(status_code=400, detail="Nao foi possivel ler a planilha do iFood") from error

@api_router.post("/admin/cash-entries")
def create_cash_entry(entry: CashEntryCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    now_iso = datetime.now().isoformat(timespec='seconds')
    data_lancamento = entry.data_lancamento or now_iso
    conn.execute(
        """
        INSERT INTO cash_entries (
            uuid, tipo, categoria, descricao, valor,
            forma_pagamento, data_lancamento, observacao, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            new_uuid,
            entry.tipo,
            entry.categoria,
            entry.descricao,
            entry.valor,
            entry.forma_pagamento,
            data_lancamento,
            entry.observacao,
            now_iso,
        ),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM cash_entries WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.put("/admin/cash-entries/{entry_uuid}")
def update_cash_entry(entry_uuid: str, entry: CashEntryUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT * FROM cash_entries WHERE uuid = ?", (entry_uuid,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Cash entry not found")
    updates = {k: v for k, v in entry.model_dump(exclude_unset=True).items()}
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(
            f"UPDATE cash_entries SET {set_clause} WHERE uuid = ?",
            list(updates.values()) + [entry_uuid],
        )
        conn.commit()
    row = conn.execute("SELECT * FROM cash_entries WHERE uuid = ?", (entry_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.delete("/admin/cash-entries/{entry_uuid}")
def delete_cash_entry(entry_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM cash_entries WHERE uuid = ?", (entry_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Cash entry deleted"}

@api_router.get("/admin/stock-items")
def get_stock_items():
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    rows = conn.execute("SELECT * FROM stock_items ORDER BY nome").fetchall()
    conn.close()
    result = []
    for row in rows:
        item = row_to_dict(row)
        quantidade = float(item.get("quantidade") or 0)
        valor_pago = float(item.get("valor_pago") or 0)
        item["unidade"] = "kg"
        item["custo_unitario"] = (valor_pago / quantidade) if quantidade > 0 else 0
        item["custo_por_grama"] = (valor_pago / (quantidade * 1000)) if quantidade > 0 else 0
        result.append(item)
    return result

@api_router.post("/admin/stock-items")
def create_stock_item(item: StockItemCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    conn.execute(
        "INSERT INTO stock_items (uuid, nome, unidade, quantidade, valor_pago, ativo) VALUES (?, ?, ?, ?, ?, ?)",
        (new_uuid, item.nome, "kg", item.quantidade, item.valor_pago, 1 if item.ativo else 0),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM stock_items WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    data = row_to_dict(row)
    quantidade = float(data.get("quantidade") or 0)
    valor_pago = float(data.get("valor_pago") or 0)
    data["unidade"] = "kg"
    data["custo_unitario"] = (valor_pago / quantidade) if quantidade > 0 else 0
    data["custo_por_grama"] = (valor_pago / (quantidade * 1000)) if quantidade > 0 else 0
    return data

@api_router.put("/admin/stock-items/{item_uuid}")
def update_stock_item(item_uuid: str, item: StockItemUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT * FROM stock_items WHERE uuid = ?", (item_uuid,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Stock item not found")
    updates = {}
    for field, value in item.model_dump(exclude_unset=True).items():
        if field == 'ativo':
            updates[field] = 1 if value else 0
        elif field == 'unidade':
            updates[field] = "kg"
        else:
            updates[field] = value
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(f"UPDATE stock_items SET {set_clause} WHERE uuid = ?", list(updates.values()) + [item_uuid])
        conn.commit()
    row = conn.execute("SELECT * FROM stock_items WHERE uuid = ?", (item_uuid,)).fetchone()
    conn.close()
    data = row_to_dict(row)
    quantidade = float(data.get("quantidade") or 0)
    valor_pago = float(data.get("valor_pago") or 0)
    data["unidade"] = "kg"
    data["custo_unitario"] = (valor_pago / quantidade) if quantidade > 0 else 0
    data["custo_por_grama"] = (valor_pago / (quantidade * 1000)) if quantidade > 0 else 0
    return data

@api_router.delete("/admin/stock-items/{item_uuid}")
def delete_stock_item(item_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM product_recipes WHERE stock_item_uuid = ?", (item_uuid,))
    conn.execute("DELETE FROM stock_items WHERE uuid = ?", (item_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Stock item deleted"}

@api_router.get("/admin/product-recipes")
def get_product_recipes():
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    rows = conn.execute(
        """
        SELECT
            pr.*,
            p.nome AS product_nome,
            si.nome AS stock_item_nome,
            si.unidade AS stock_item_unidade,
            si.quantidade AS stock_item_quantidade,
            si.valor_pago AS stock_item_valor_pago
        FROM product_recipes pr
        JOIN products p ON p.uuid = pr.product_uuid
        JOIN stock_items si ON si.uuid = pr.stock_item_uuid
        ORDER BY p.nome, si.nome
        """
    ).fetchall()
    conn.close()
    result = []
    for row in rows:
        data = row_to_dict(row)
        quantidade_estoque = float(data.get("stock_item_quantidade") or 0)
        valor_pago = float(data.get("stock_item_valor_pago") or 0)
        custo_unitario = (valor_pago / quantidade_estoque) if quantidade_estoque > 0 else 0
        custo_por_grama = (valor_pago / (quantidade_estoque * 1000)) if quantidade_estoque > 0 else 0
        data["stock_item_unidade"] = "kg"
        data["custo_unitario_insumo"] = custo_unitario
        data["custo_por_grama_insumo"] = custo_por_grama
        data["custo_estimado"] = custo_por_grama * float(data.get("quantidade_utilizada") or 0)
        result.append(data)
    return result

@api_router.post("/admin/product-recipes")
def create_product_recipe(recipe: ProductRecipeCreate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    new_uuid = str(uuid_lib.uuid4())
    conn.execute(
        "INSERT INTO product_recipes (uuid, product_uuid, stock_item_uuid, quantidade_utilizada) VALUES (?, ?, ?, ?)",
        (new_uuid, recipe.product_uuid, recipe.stock_item_uuid, recipe.quantidade_utilizada),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM product_recipes WHERE uuid = ?", (new_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.put("/admin/product-recipes/{recipe_uuid}")
def update_product_recipe(recipe_uuid: str, recipe: ProductRecipeUpdate):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    existing = conn.execute("SELECT * FROM product_recipes WHERE uuid = ?", (recipe_uuid,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Product recipe not found")
    updates = {k: v for k, v in recipe.model_dump(exclude_unset=True).items()}
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(f"UPDATE product_recipes SET {set_clause} WHERE uuid = ?", list(updates.values()) + [recipe_uuid])
        conn.commit()
    row = conn.execute("SELECT * FROM product_recipes WHERE uuid = ?", (recipe_uuid,)).fetchone()
    conn.close()
    return row_to_dict(row)

@api_router.delete("/admin/product-recipes/{recipe_uuid}")
def delete_product_recipe(recipe_uuid: str):
    if not IS_DEVELOPMENT:
        raise HTTPException(status_code=403, detail="Admin not available in production")
    conn = get_db()
    conn.execute("DELETE FROM product_recipes WHERE uuid = ?", (recipe_uuid,))
    conn.commit()
    conn.close()
    return {"message": "Product recipe deleted"}

app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

init_db()
logger.info(f"Database initialized at {DB_PATH}")
logger.info(f"Development mode: {IS_DEVELOPMENT}")
